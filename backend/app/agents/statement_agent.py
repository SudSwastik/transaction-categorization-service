import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

# Matches the numeric portion of an amount cell after commas are stripped, e.g.
# "$1234.56" -> "1234.56", "USD 1234.56" -> "1234.56". Sign is handled separately
# since bank exports mark negatives with parentheses or a trailing "-" as often
# as a leading one.
_AMOUNT_NUMERIC_RE = re.compile(r"-?\d[\d,]*\.?\d*")

# Header keywords are matched as case-insensitive substrings against each cell in
# a candidate header row — statements from different banks label the same column
# differently ("Narration" vs "Description", "Txn Date" vs "Value Date").
_COLUMN_KEYWORDS: dict[str, tuple[str, ...]] = {
    "date": ("transaction date", "txn date", "value date", "posted date", "date"),
    "description": ("description", "narration", "particulars", "details", "memo"),
    "debit": ("debit", "withdrawal", "dr"),
    "credit": ("credit", "deposit", "cr"),
    "amount": ("transaction amount", "txn amount", "amount"),
    "balance": ("running balance", "closing balance", "balance"),
}
_HEADER_SCAN_ROWS = 10


class StatementParseError(Exception):
    pass


@dataclass(frozen=True, slots=True)
class RawTransaction:
    raw_description: str
    transaction_date: date
    raw_amount: Decimal


class StatementAgent:
    """Parses raw statement files (currently XLSX) into `RawTransaction` rows.

    Heuristic, not schema-bound: bank export formats vary in column order,
    naming, and whether debit/credit are split or signed into one column.
    """

    def parse_xlsx(self, content: bytes) -> list[RawTransaction]:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        if sheet is None:
            raise StatementParseError("Workbook has no active sheet")

        rows = list(sheet.iter_rows(values_only=True))
        header_row_index, columns = self._detect_columns(rows[:_HEADER_SCAN_ROWS])

        transactions: list[RawTransaction] = []
        for row in rows[header_row_index + 1 :]:
            if row is None or all(cell is None for cell in row):
                continue

            raw_date = row[columns["date"]] if "date" in columns else None
            raw_description = row[columns["description"]] if "description" in columns else None
            if raw_date is None or raw_description is None:
                continue

            transaction_date = self._coerce_date(raw_date)
            if transaction_date is None:
                continue

            amount = self._extract_amount(row, columns)
            if amount is None:
                continue

            transactions.append(
                RawTransaction(
                    raw_description=str(raw_description).strip(),
                    transaction_date=transaction_date,
                    raw_amount=amount,
                )
            )

        return transactions

    def _detect_columns(
        self, candidate_rows: list[tuple[object, ...] | None]
    ) -> tuple[int, dict[str, int]]:
        """Finds the header row among the first few rows and maps each recognized
        column to its semantic role. Picks the row with the most keyword hits,
        requiring at minimum a date and an amount-bearing column (single amount,
        or a debit/credit pair)."""
        best_index = -1
        best_columns: dict[str, int] = {}
        best_score = 0

        for row_index, row in enumerate(candidate_rows):
            if row is None:
                continue
            columns = self._match_row_to_columns(row)
            has_amount_source = "amount" in columns or "debit" in columns or "credit" in columns
            if "date" not in columns or "description" not in columns or not has_amount_source:
                continue
            score = len(columns)
            if score > best_score:
                best_score = score
                best_index = row_index
                best_columns = columns

        if best_index == -1:
            raise StatementParseError(
                "Could not detect a header row with date/description/amount columns"
            )
        return best_index, best_columns

    def _match_row_to_columns(self, row: tuple[object, ...]) -> dict[str, int]:
        columns: dict[str, int] = {}
        for cell_index, cell in enumerate(row):
            if not isinstance(cell, str):
                continue
            cell_lower = cell.strip().lower()
            for role, keywords in _COLUMN_KEYWORDS.items():
                if role in columns:
                    continue
                if any(keyword in cell_lower for keyword in keywords):
                    columns[role] = cell_index
                    break
        return columns

    def _extract_amount(self, row: tuple[object, ...], columns: dict[str, int]) -> Decimal | None:
        if "amount" in columns:
            return self._parse_amount(row[columns["amount"]])

        debit = self._parse_amount(row[columns["debit"]]) if "debit" in columns else None
        credit = self._parse_amount(row[columns["credit"]]) if "credit" in columns else None
        if debit:
            return -abs(debit)
        if credit:
            return abs(credit)
        return None

    def _parse_amount(self, value: object) -> Decimal | None:
        """Parses an amount cell that may be a raw number or bank-formatted text:
        parenthesized negatives `(1,234.56)`, trailing-minus negatives `1234.56-`,
        thousands separators, and currency prefixes/suffixes (`$`, `USD`, ...)."""
        if value is None or value == "":
            return None
        if isinstance(value, int | float | Decimal):
            return Decimal(str(value))

        text = str(value).strip()
        if not text:
            return None

        negative = False
        if text.startswith("(") and text.endswith(")"):
            negative = True
            text = text[1:-1]
        elif text.endswith("-"):
            negative = True
            text = text[:-1]

        match = _AMOUNT_NUMERIC_RE.search(text.replace(",", ""))
        if match is None:
            return None

        try:
            amount = Decimal(match.group())
        except InvalidOperation:
            return None

        return -abs(amount) if negative else amount

    def _coerce_date(self, value: object) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    continue
        return None
